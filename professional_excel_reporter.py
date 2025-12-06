"""
Professional Excel Reporter for Fund Managers
Generates comprehensive daily transaction logs, P&L tracking, and performance metrics.
"""
import os
import pandas as pd
import numpy as np
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import quantstats as qs


class ProfessionalExcelReporter:
    """Creates professional-grade Excel reports for fund managers."""
    
    def __init__(self, initial_capital: float, report_dir: str = "reports/baremetal_optimization"):
        """
        Initialize the reporter.
        
        Args:
            initial_capital: Starting portfolio value
            report_dir: Directory to save reports
        """
        self.initial_capital = initial_capital
        self.report_dir = report_dir
        self.transactions = []
        self.daily_positions = []
        
    def log_rebalance(self, date, old_weights: pd.Series, new_weights: pd.Series, 
                     prices: pd.Series, portfolio_value: float, reason: str = "Rebalancing",
                     ib_execution_info: dict = None):
        """
        Log a rebalancing event with transaction details.
        
        Args:
            date: Rebalancing date
            old_weights: Previous weights
            new_weights: New target weights
            prices: Current prices for each ETF
            portfolio_value: Current portfolio value
            reason: Reason for the transaction (e.g., "Emergency Exit", "Re-entry", "Scheduled Rebalancing")
            ib_execution_info: Optional IB execution details dict with:
                - fill_ratios: pd.Series of fill ratios per ETF
                - execution_prices: pd.Series of actual execution prices (with slippage)
                - execution_delay: int (days until execution, 0 = same day, 1 = next day)
                - order_type: str (e.g., "MARKET_ON_OPEN", "MARKET")
                - execution_cost: float (total execution cost in dollars)
        """
        # Extract IB execution details if provided
        fill_ratios = ib_execution_info.get('fill_ratios', None) if ib_execution_info else None
        execution_prices = ib_execution_info.get('execution_prices', None) if ib_execution_info else None
        execution_delay = ib_execution_info.get('execution_delay', 0) if ib_execution_info else 0
        order_type = ib_execution_info.get('order_type', None) if ib_execution_info else None
        execution_cost = ib_execution_info.get('execution_cost', 0.0) if ib_execution_info else 0.0
        
        for etf in new_weights.index:
            old_weight = old_weights.get(etf, 0.0)
            new_weight = new_weights.get(etf, 0.0)
            
            if abs(new_weight - old_weight) > 0.0001:  # Only log meaningful changes
                old_value = portfolio_value * old_weight
                new_value = portfolio_value * new_weight
                change_value = new_value - old_value
                
                # Use execution price if available (includes slippage), otherwise use market price
                if execution_prices is not None and etf in execution_prices.index:
                    execution_price = execution_prices[etf]
                    if pd.isna(execution_price) or execution_price <= 0:
                        execution_price = prices.get(etf, np.nan)
                else:
                    execution_price = prices.get(etf, np.nan)
                
                price = execution_price  # Use execution price (with slippage) for reporting
                market_price = prices.get(etf, np.nan)  # Also keep market price for comparison
                
                old_shares = old_value / price if price > 0 and not np.isnan(price) else 0
                new_shares = new_value / price if price > 0 and not np.isnan(price) else 0
                shares_traded = new_shares - old_shares
                
                # Calculate slippage if we have both market and execution prices
                slippage = 0.0
                slippage_pct = 0.0
                if not pd.isna(market_price) and not pd.isna(price) and market_price > 0:
                    slippage = price - market_price  # Positive = worse execution (for sells)
                    slippage_pct = (slippage / market_price) * 100 if market_price > 0 else 0.0
                
                # Get fill ratio if available
                fill_ratio = fill_ratios.get(etf, 1.0) if fill_ratios is not None else 1.0
                
                transaction_record = {
                    'Date': date,
                    'ETF': etf,
                    'Action': 'BUY' if shares_traded > 0 else 'SELL',
                    'Shares': abs(shares_traded),
                    'Market_Price': market_price,
                    'Execution_Price': price,
                    'Slippage': slippage,
                    'Slippage_Pct': slippage_pct,
                    'Fill_Ratio': fill_ratio * 100,  # Convert to percentage
                    'Value': abs(change_value),
                    'Old_Weight': old_weight,
                    'New_Weight': new_weight,
                    'Weight_Change': new_weight - old_weight,
                    'Reason': reason
                }
                
                # Add IB execution metadata if available
                if ib_execution_info:
                    transaction_record['Execution_Delay_Days'] = execution_delay
                    transaction_record['Order_Type'] = order_type.value if hasattr(order_type, 'value') else str(order_type)
                    # Execution cost is per transaction, distribute proportionally
                    if abs(change_value) > 0:
                        total_value = sum(abs(portfolio_value * (new_weights.get(e, 0.0) - old_weights.get(e, 0.0))) 
                                        for e in new_weights.index)
                        if total_value > 0:
                            transaction_record['Execution_Cost'] = execution_cost * (abs(change_value) / total_value)
                        else:
                            transaction_record['Execution_Cost'] = 0.0
                    else:
                        transaction_record['Execution_Cost'] = 0.0
                
                self.transactions.append(transaction_record)
    
    def log_daily_position(self, date, weights: pd.Series, prices: pd.Series, 
                          portfolio_value: float, daily_returns: pd.Series = None,
                          transaction_cost: float = 0.0, cumulative_transaction_costs: float = 0.0):
        """
        Log daily position data for each ETF.
        
        Args:
            date: Trading date
            weights: Current weights for each ETF
            prices: Current prices for each ETF
            portfolio_value: Current portfolio value
            daily_returns: Daily returns for each ETF (optional)
            transaction_cost: Transaction cost for this day (optional)
            cumulative_transaction_costs: Cumulative transaction costs up to this day (optional)
        """
        position_data = {
            'Date': date,
            'Portfolio_Value': portfolio_value,
            'Portfolio_Return': 0.0,  # Will be calculated later
            'Transaction_Cost': transaction_cost,
            'Cumulative_Transaction_Costs': cumulative_transaction_costs
        }
        
        for etf in weights.index:
            weight = weights.get(etf, 0.0)
            price = prices.get(etf, np.nan)
            position_value = portfolio_value * weight
            shares = position_value / price if price > 0 and not np.isnan(price) else 0
            daily_return = daily_returns.get(etf, 0.0) if daily_returns is not None else 0.0
            position_pnl = position_value * daily_return
            
            position_data[f'{etf}_Weight'] = weight
            position_data[f'{etf}_Price'] = price
            position_data[f'{etf}_Shares'] = shares
            position_data[f'{etf}_Value'] = position_value
            position_data[f'{etf}_DailyReturn'] = daily_return
            position_data[f'{etf}_DailyPnL'] = position_pnl
        
        self.daily_positions.append(position_data)
    
    def generate_comprehensive_report(self, optimizer_equity: pd.Series, 
                                     benchmark_equity: pd.Series,
                                     etf_list: list,
                                     cumulative_transaction_costs: float = 0.0,
                                     ib_execution_summary: pd.DataFrame = None) -> str:
        """
        Generate a comprehensive Excel report with multiple sheets.
        
        Args:
            optimizer_equity: Portfolio equity curve
            benchmark_equity: Benchmark equity curve
            etf_list: List of ETF tickers
            cumulative_transaction_costs: Total transaction costs (optional)
            
        Returns:
            Path to the generated Excel file
        """
        print("\n" + "=" * 80)
        print("GENERATING PROFESSIONAL EXCEL REPORT")
        print("=" * 80)
        
        # Create DataFrames from logged data
        transactions_df = pd.DataFrame(self.transactions) if self.transactions else pd.DataFrame()
        positions_df = pd.DataFrame(self.daily_positions) if self.daily_positions else pd.DataFrame()
        
        # Prepare report filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"professional_fund_report_{timestamp}.xlsx"
        filepath = os.path.join(self.report_dir, filename)
        os.makedirs(self.report_dir, exist_ok=True)
        
        # Create Excel writer
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            # Sheet 1: Executive Summary
            print("  Creating Executive Summary...")
            self._create_executive_summary(writer, optimizer_equity, benchmark_equity, cumulative_transaction_costs)
            
            # Sheet 2: Daily Portfolio Performance
            print("  Creating Daily Portfolio Performance...")
            self._create_daily_performance(writer, positions_df, etf_list)
            
            # Sheet 3: Transaction Log
            print("  Creating Transaction Log...")
            self._create_transaction_log(writer, transactions_df)
            
            # Sheet 4: Daily Holdings Detail
            print("  Creating Daily Holdings Detail...")
            self._create_daily_holdings(writer, positions_df, etf_list)
            
            # Sheet 5: ETF Performance Matrix
            print("  Creating ETF Performance Matrix...")
            self._create_etf_performance_matrix(writer, positions_df, etf_list)
            
            # Sheet 6: Monthly Summary
            print("  Creating Monthly Summary...")
            self._create_monthly_summary(writer, positions_df)
            
            # Sheet 7: Risk Metrics
            print("  Creating Risk Metrics...")
            self._create_risk_metrics(writer, positions_df, optimizer_equity)
            
            # Sheet 8: IB Execution Summary (always create if IB execution is enabled, even if empty)
            if ib_execution_summary is not None:
                print("  Creating IB Execution Summary...")
                if ib_execution_summary.empty:
                    # Create empty summary with message
                    empty_msg = pd.DataFrame({
                        'Message': [
                            'IB Execution Model is enabled but no trades were executed through it.',
                            'Trades are currently using simplified execution (instant, perfect fills).',
                            'To use IB execution, modify the code to call ib_execution.execute_trade() when trades occur.',
                            '',
                            'IB Execution features:',
                            '- EOD signals → Next day execution (1-day delay)',
                            '- Partial fills: 50-85% for large orders',
                            '- Slippage: 0.03-0.30% (depending on order size)',
                            '- Price impact: 0.01% per 1% of portfolio traded'
                        ]
                    })
                    self._create_ib_execution_summary(writer, empty_msg)
                else:
                    self._create_ib_execution_summary(writer, ib_execution_summary)
        
        # Apply formatting
        print("  Applying professional formatting...")
        self._apply_workbook_formatting(filepath)
        
        print(f"\n[OK] Professional Excel report saved: {filepath}")
        print("=" * 80)
        return filepath
    
    def _create_executive_summary(self, writer, optimizer_equity: pd.Series, 
                                 benchmark_equity: pd.Series, 
                                 cumulative_transaction_costs: float = 0.0):
        """Create executive summary sheet."""
        # Ensure DatetimeIndex for QuantStats
        if not isinstance(optimizer_equity.index, pd.DatetimeIndex):
            optimizer_equity.index = pd.to_datetime(optimizer_equity.index)
        if not isinstance(benchmark_equity.index, pd.DatetimeIndex):
            benchmark_equity.index = pd.to_datetime(benchmark_equity.index)
        
        optimizer_returns = optimizer_equity.pct_change().dropna()
        benchmark_returns = benchmark_equity.pct_change().dropna()
        
        # Use QuantStats for all metrics
        summary_data = {
            'Metric': [
                'Initial Capital',
                'Final Portfolio Value',
                'Total Return',
                'CAGR',
                'Total Days',
                'Sharpe Ratio',
                'Sortino Ratio',
                'Max Drawdown',
                'Volatility (Annual)',
                'Best Day',
                'Worst Day',
                'Winning Days %',
                '',
                'Total Transaction Costs',
                'Transaction Costs as % of Initial Capital',
                'Transaction Costs as % of Final Value',
                '',
                'Benchmark Final Value',
                'Benchmark Total Return',
                'Benchmark CAGR',
                'Benchmark Sharpe Ratio',
                'Benchmark Max Drawdown',
                '',
                'Alpha vs Benchmark',
                'Excess Return vs Benchmark'
            ],
            'Value': [
                self.initial_capital,
                optimizer_equity.iloc[-1],
                (optimizer_equity.iloc[-1] / self.initial_capital - 1),
                qs.stats.cagr(optimizer_returns),
                len(optimizer_equity),
                qs.stats.sharpe(optimizer_returns),
                qs.stats.sortino(optimizer_returns),
                qs.stats.max_drawdown(optimizer_returns),
                qs.stats.volatility(optimizer_returns),
                optimizer_returns.max(),
                optimizer_returns.min(),
                qs.stats.win_rate(optimizer_returns),
                '',
                cumulative_transaction_costs,
                (cumulative_transaction_costs / self.initial_capital) if self.initial_capital > 0 else 0.0,
                (cumulative_transaction_costs / optimizer_equity.iloc[-1]) if optimizer_equity.iloc[-1] > 0 else 0.0,
                '',
                benchmark_equity.iloc[-1],
                (benchmark_equity.iloc[-1] / self.initial_capital - 1),
                qs.stats.cagr(benchmark_returns),
                qs.stats.sharpe(benchmark_returns),
                qs.stats.max_drawdown(benchmark_returns),
                '',
                qs.stats.cagr(optimizer_returns) - qs.stats.cagr(benchmark_returns),
                (optimizer_equity.iloc[-1] - benchmark_equity.iloc[-1])
            ]
        }
        
        summary_df = pd.DataFrame(summary_data)
        summary_df.to_excel(writer, sheet_name='Executive Summary', index=False)
    
    def _create_daily_performance(self, writer, positions_df: pd.DataFrame, etf_list: list):
        """Create daily portfolio performance sheet."""
        if positions_df.empty:
            pd.DataFrame({'Message': ['No data available']}).to_excel(
                writer, sheet_name='Daily Performance', index=False)
            return
        
        # Calculate daily P&L
        positions_df = positions_df.copy()
        positions_df['Daily_PnL'] = positions_df['Portfolio_Value'].diff()
        positions_df['Daily_Return'] = positions_df['Portfolio_Value'].pct_change()
        positions_df['Cumulative_PnL'] = positions_df['Portfolio_Value'] - self.initial_capital
        positions_df['Cumulative_Return'] = (positions_df['Portfolio_Value'] / self.initial_capital) - 1
        
        # Calculate cumulative returns
        positions_df['Cumulative_Return_Pct'] = positions_df['Cumulative_Return'] * 100
        
        # Add transaction cost columns if they exist
        display_cols = [
            'Date', 'Portfolio_Value', 'Daily_PnL', 'Daily_Return', 
            'Cumulative_PnL', 'Cumulative_Return'
        ]
        
        # Add transaction cost columns if available
        if 'Transaction_Cost' in positions_df.columns:
            display_cols.append('Transaction_Cost')
        if 'Cumulative_Transaction_Costs' in positions_df.columns:
            display_cols.append('Cumulative_Transaction_Costs')
        
        # Only include columns that exist
        available_cols = [col for col in display_cols if col in positions_df.columns]
        performance_df = positions_df[available_cols].copy()
        performance_df.to_excel(writer, sheet_name='Daily Performance', index=False)
    
    def _create_transaction_log(self, writer, transactions_df: pd.DataFrame):
        """Create transaction log sheet."""
        if transactions_df.empty:
            pd.DataFrame({'Message': ['No transactions recorded']}).to_excel(
                writer, sheet_name='Transaction Log', index=False)
            return
        
        # Format and sort transactions
        transactions_df = transactions_df.sort_values('Date').copy()
        transactions_df['Date'] = pd.to_datetime(transactions_df['Date'])
        
        # Reorder columns for better presentation
        # Check which IB execution columns exist
        base_cols = ['Date', 'ETF', 'Action', 'Reason']
        ib_cols = ['Execution_Delay_Days', 'Order_Type', 'Fill_Ratio', 'Market_Price', 
                   'Execution_Price', 'Slippage', 'Slippage_Pct', 'Execution_Cost']
        standard_cols = ['Shares', 'Price', 'Value', 'Old_Weight', 'New_Weight', 'Weight_Change']
        
        # Build column order: base columns, then IB columns (if exist), then standard columns
        col_order = base_cols.copy()
        for col in ib_cols:
            if col in transactions_df.columns:
                col_order.append(col)
        for col in standard_cols:
            if col in transactions_df.columns:
                col_order.append(col)
        
        # Only include columns that actually exist
        col_order = [col for col in col_order if col in transactions_df.columns]
        transactions_df = transactions_df[col_order]
        
        transactions_df.to_excel(writer, sheet_name='Transaction Log', index=False)
    
    def _create_daily_holdings(self, writer, positions_df: pd.DataFrame, etf_list: list):
        """Create detailed daily holdings sheet."""
        if positions_df.empty:
            pd.DataFrame({'Message': ['No data available']}).to_excel(
                writer, sheet_name='Daily Holdings', index=False)
            return
        
        # Prepare holdings data with each ETF as columns
        holdings_data = {'Date': positions_df['Date'], 
                        'Portfolio_Value': positions_df['Portfolio_Value']}
        
        for etf in etf_list:
            if f'{etf}_Weight' in positions_df.columns:
                holdings_data[f'{etf}_Weight'] = positions_df[f'{etf}_Weight']
                holdings_data[f'{etf}_Value'] = positions_df[f'{etf}_Value']
                holdings_data[f'{etf}_Shares'] = positions_df[f'{etf}_Shares']
        
        holdings_df = pd.DataFrame(holdings_data)
        holdings_df.to_excel(writer, sheet_name='Daily Holdings', index=False)
    
    def _create_etf_performance_matrix(self, writer, positions_df: pd.DataFrame, etf_list: list):
        """Create ETF performance comparison matrix."""
        if positions_df.empty:
            pd.DataFrame({'Message': ['No data available']}).to_excel(
                writer, sheet_name='ETF Performance', index=False)
            return
        
        performance_metrics = []
        
        for etf in etf_list:
            if f'{etf}_DailyPnL' in positions_df.columns:
                daily_pnl = positions_df[f'{etf}_DailyPnL']
                weights = positions_df[f'{etf}_Weight']
                values = positions_df[f'{etf}_Value']
                
                total_pnl = daily_pnl.sum()
                avg_weight = weights.mean()
                max_weight = weights.max()
                avg_value = values.mean()
                days_held = (weights > 0.001).sum()
                
                performance_metrics.append({
                    'ETF': etf,
                    'Total_PnL': total_pnl,
                    'Avg_Weight': avg_weight,
                    'Max_Weight': max_weight,
                    'Avg_Position_Value': avg_value,
                    'Days_Held': days_held,
                    'Contribution_%': 0  # Will calculate after loop
                })
        
        performance_df = pd.DataFrame(performance_metrics)
        
        if not performance_df.empty:
            total_pnl = performance_df['Total_PnL'].sum()
            if total_pnl != 0:
                performance_df['Contribution_%'] = (performance_df['Total_PnL'] / total_pnl) * 100
            
            performance_df = performance_df.sort_values('Total_PnL', ascending=False)
        
        performance_df.to_excel(writer, sheet_name='ETF Performance', index=False)
    
    def _create_monthly_summary(self, writer, positions_df: pd.DataFrame):
        """Create monthly summary sheet."""
        if positions_df.empty:
            pd.DataFrame({'Message': ['No data available']}).to_excel(
                writer, sheet_name='Monthly Summary', index=False)
            return
        
        positions_df = positions_df.copy()
        positions_df['Date'] = pd.to_datetime(positions_df['Date'])
        positions_df['YearMonth'] = positions_df['Date'].dt.to_period('M')
        
        # Group by month
        monthly_data = []
        for period, group in positions_df.groupby('YearMonth'):
            start_value = group['Portfolio_Value'].iloc[0]
            end_value = group['Portfolio_Value'].iloc[-1]
            monthly_return = (end_value / start_value - 1) if start_value > 0 else 0
            
            monthly_data.append({
                'Month': str(period),
                'Start_Value': start_value,
                'End_Value': end_value,
                'Monthly_Return': monthly_return,
                'Trading_Days': len(group),
                'Best_Day': group['Portfolio_Value'].pct_change().max(),
                'Worst_Day': group['Portfolio_Value'].pct_change().min()
            })
        
        monthly_df = pd.DataFrame(monthly_data)
        monthly_df.to_excel(writer, sheet_name='Monthly Summary', index=False)
    
    def _create_ib_execution_summary(self, writer, ib_execution_summary: pd.DataFrame):
        """Create IB execution summary sheet with execution statistics."""
        # Check if this is an empty message DataFrame (from our empty case)
        if ib_execution_summary.empty or ('Message' in ib_execution_summary.columns and len(ib_execution_summary) > 0):
            # If it has a Message column, write it directly
            if 'Message' in ib_execution_summary.columns:
                ib_execution_summary.to_excel(writer, sheet_name='IB Execution Summary', index=False)
                return
            else:
                # Truly empty - create message
                pd.DataFrame({'Message': ['No IB execution data available - IB Execution Model is enabled but no trades were executed through it.']}).to_excel(
                    writer, sheet_name='IB Execution Summary', index=False)
                return
        
        # Format the summary
        summary_df = ib_execution_summary.copy()
        
        # Calculate summary statistics
        stats_data = []
        if 'execution_cost' in summary_df.columns:
            stats_data.append({'Metric': 'Total_Executions', 'Value': len(summary_df)})
            stats_data.append({'Metric': 'Total_Execution_Cost', 'Value': summary_df['execution_cost'].sum()})
            stats_data.append({'Metric': 'Avg_Execution_Cost', 'Value': summary_df['execution_cost'].mean()})
        if 'avg_fill_ratio' in summary_df.columns:
            stats_data.append({'Metric': 'Avg_Fill_Ratio_%', 'Value': summary_df['avg_fill_ratio'].mean() * 100})
        if 'min_fill_ratio' in summary_df.columns:
            stats_data.append({'Metric': 'Min_Fill_Ratio_%', 'Value': summary_df['min_fill_ratio'].min() * 100})
        if 'execution_delay' in summary_df.columns:
            stats_data.append({'Metric': 'Avg_Execution_Delay_Days', 'Value': summary_df['execution_delay'].mean()})
        if 'total_unfilled_pct' in summary_df.columns:
            stats_data.append({'Metric': 'Total_Unfilled_%', 'Value': summary_df['total_unfilled_pct'].sum()})
        if 'order_type' in summary_df.columns:
            stats_data.append({'Metric': 'Market_On_Open_Orders', 
                             'Value': len(summary_df[summary_df['order_type'] == 'MARKET_ON_OPEN'])})
            stats_data.append({'Metric': 'Market_Orders', 
                             'Value': len(summary_df[summary_df['order_type'] == 'MARKET'])})
        if 'order_direction' in summary_df.columns:
            stats_data.append({'Metric': 'SELL_Orders', 
                             'Value': len(summary_df[summary_df['order_direction'] == 'SELL'])})
            stats_data.append({'Metric': 'BUY_Orders', 
                             'Value': len(summary_df[summary_df['order_direction'] == 'BUY'])})
        
        stats_df = pd.DataFrame(stats_data)
        
        # Write summary statistics first, then detailed data
        stats_df.to_excel(writer, sheet_name='IB Execution Summary', index=False, startrow=0)
        
        # Add detailed summary below (with some spacing)
        if not summary_df.empty:
            summary_df.to_excel(writer, sheet_name='IB Execution Summary', index=False, startrow=len(stats_df) + 3)
    
    def _create_risk_metrics(self, writer, positions_df: pd.DataFrame, 
                            optimizer_equity: pd.Series):
        """Create risk metrics sheet."""
        if positions_df.empty:
            pd.DataFrame({'Message': ['No data available']}).to_excel(
                writer, sheet_name='Risk Metrics', index=False)
            return
        
        returns = optimizer_equity.pct_change().dropna()
        
        # Calculate rolling metrics
        rolling_windows = [21, 63, 126, 252]  # 1M, 3M, 6M, 1Y
        risk_data = []
        
        for window in rolling_windows:
            if len(returns) >= window:
                rolling_vol = returns.rolling(window).std() * np.sqrt(252)
                rolling_sharpe = (returns.rolling(window).mean() * 252) / (returns.rolling(window).std() * np.sqrt(252))
                
                risk_data.append({
                    'Period': f'{window} Days',
                    'Avg_Volatility': rolling_vol.mean(),
                    'Max_Volatility': rolling_vol.max(),
                    'Min_Volatility': rolling_vol.min(),
                    'Avg_Sharpe': rolling_sharpe.mean(),
                    'Max_Sharpe': rolling_sharpe.max(),
                    'Min_Sharpe': rolling_sharpe.min()
                })
        
        # Value at Risk (VaR)
        var_95 = returns.quantile(0.05)
        var_99 = returns.quantile(0.01)
        cvar_95 = returns[returns <= var_95].mean()
        
        risk_data.append({
            'Period': 'VaR (95%)',
            'Avg_Volatility': var_95,
            'Max_Volatility': np.nan,
            'Min_Volatility': np.nan,
            'Avg_Sharpe': np.nan,
            'Max_Sharpe': np.nan,
            'Min_Sharpe': np.nan
        })
        
        risk_data.append({
            'Period': 'VaR (99%)',
            'Avg_Volatility': var_99,
            'Max_Volatility': np.nan,
            'Min_Volatility': np.nan,
            'Avg_Sharpe': np.nan,
            'Max_Sharpe': np.nan,
            'Min_Sharpe': np.nan
        })
        
        risk_data.append({
            'Period': 'CVaR (95%)',
            'Avg_Volatility': cvar_95,
            'Max_Volatility': np.nan,
            'Min_Volatility': np.nan,
            'Avg_Sharpe': np.nan,
            'Max_Sharpe': np.nan,
            'Min_Sharpe': np.nan
        })
        
        risk_df = pd.DataFrame(risk_data)
        risk_df.to_excel(writer, sheet_name='Risk Metrics', index=False)
    
    def _apply_workbook_formatting(self, filepath: str):
        """Apply professional formatting to the workbook."""
        from openpyxl import load_workbook
        
        wb = load_workbook(filepath)
        
        # Define styles
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True, size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Format header row
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
            
            # Special formatting for Executive Summary
            if sheet_name == 'Executive Summary':
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row, 
                                       min_col=1, max_col=ws.max_column):
                    for cell in row:
                        cell.border = border
                        cell.alignment = Alignment(vertical='center')
                        
                        # Format based on metric name in column A
                        if cell.column == 2 and cell.value is not None and isinstance(cell.value, (int, float)):
                            metric_name = ws.cell(cell.row, 1).value
                            if metric_name and isinstance(metric_name, str):
                                metric_lower = metric_name.lower()
                                
                                # Currency metrics (must check first before percentage)
                                if any(keyword in metric_lower for keyword in ['capital', 'value', 'excess return']):
                                    cell.number_format = '$#,##0.00'
                                # Days metric (integer)
                                elif 'days' in metric_lower:
                                    cell.number_format = '#,##0'
                                # Ratio metrics (no formatting, just 2 decimals)
                                elif any(keyword in metric_lower for keyword in ['sharpe', 'sortino']):
                                    cell.number_format = '0.00'
                                # Percentage metrics
                                elif any(keyword in metric_lower for keyword in ['return', 'cagr', 
                                                                               'drawdown', 'volatility', 'alpha',
                                                                               'winning days', 'best day', 'worst day']):
                                    cell.number_format = '0.00%'
                                else:
                                    cell.number_format = '0.00'
            
            # Format other sheets
            else:
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row, 
                                       min_col=1, max_col=ws.max_column):
                    for cell in row:
                        cell.border = border
                        cell.alignment = Alignment(vertical='center')
                        
                        # Format numbers
                        if cell.value is not None and isinstance(cell.value, (int, float)):
                            if 'Weight' in str(ws.cell(1, cell.column).value):
                                cell.number_format = '0.00%'
                            elif 'Return' in str(ws.cell(1, cell.column).value):
                                cell.number_format = '0.00%'
                            elif 'Value' in str(ws.cell(1, cell.column).value) or 'PnL' in str(ws.cell(1, cell.column).value):
                                cell.number_format = '$#,##0.00'
                            elif 'Price' in str(ws.cell(1, cell.column).value):
                                cell.number_format = '$0.00'
                            elif 'Shares' in str(ws.cell(1, cell.column).value):
                                cell.number_format = '#,##0.00'
                            else:
                                cell.number_format = '0.00'
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except (TypeError, AttributeError):
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Freeze top row
            ws.freeze_panes = 'A2'
        
        wb.save(filepath)
    
    # Helper calculation methods
    def _calculate_cagr(self, equity: pd.Series) -> float:
        """Calculate Compound Annual Growth Rate."""
        if len(equity) < 2:
            return 0.0
        total_return = equity.iloc[-1] / equity.iloc[0] - 1
        years = len(equity) / 252
        return (1 + total_return) ** (1 / years) - 1 if years > 0 else 0.0
    
    def _calculate_sharpe(self, returns: pd.Series, rf: float = 0.02) -> float:
        """Calculate Sharpe ratio."""
        if len(returns) < 2 or returns.std() == 0:
            return 0.0
        excess_returns = returns - (rf / 252)
        return np.sqrt(252) * excess_returns.mean() / returns.std()
    
    def _calculate_sortino(self, returns: pd.Series, rf: float = 0.02) -> float:
        """Calculate Sortino ratio."""
        if len(returns) < 2:
            return 0.0
        excess_returns = returns - (rf / 252)
        downside_returns = returns[returns < 0]
        if len(downside_returns) == 0 or downside_returns.std() == 0:
            return 0.0
        return np.sqrt(252) * excess_returns.mean() / downside_returns.std()
    
    def _calculate_max_drawdown(self, equity: pd.Series) -> float:
        """Calculate maximum drawdown."""
        if len(equity) < 2:
            return 0.0
        cummax = equity.expanding().max()
        drawdown = (equity - cummax) / cummax
        return drawdown.min()

